from __future__ import annotations

import csv as csv_mod
import io
import re
import uuid
from datetime import date
from typing import Any

import pyarrow as pa
import pyarrow.parquet as pq
import structlog

from access_iq.ingestion.idempotency import should_skip_if_already_successful
from access_iq.ingestion.manifests import (
    Manifest,
    ManifestStatus,
    build_manifest_prefix,
    s3_kms_args,
    utc_now_iso,
    write_manifest,
)

log = structlog.get_logger(__name__)

_EXPORT_DATE_RE = re.compile(r"export_date=(\d{8})/?$")


def _csv_bytes_to_parquet_buffer(raw_bytes: bytes) -> io.BytesIO:
    """Convert CSV bytes to Parquet buffer."""
    text = raw_bytes.decode("utf-8", errors="replace")
    reader = csv_mod.DictReader(io.StringIO(text, newline=""))
    rows = list(reader)
    if not rows:
        columns = reader.fieldnames or []
        tbl = pa.Table.from_pydict({col: [] for col in columns})
    else:
        tbl = pa.Table.from_pydict({col: [row[col] for row in rows] for col in rows[0].keys()})
    buf = io.BytesIO()
    pq.write_table(tbl, buf, compression="snappy")
    buf.seek(0)
    return buf


def _xlsx_bytes_to_parquet_buffer(raw_bytes: bytes) -> io.BytesIO:
    """Convert Excel (.xlsx) bytes to Parquet buffer."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        tbl = pa.table({})
    else:
        headers = [str(h) for h in rows[0]]
        data = {
            h: [str(row[i]) if row[i] is not None else None for row in rows[1:]]
            for i, h in enumerate(headers)
        }
        tbl = pa.Table.from_pydict(data)
    buf = io.BytesIO()
    pq.write_table(tbl, buf, compression="snappy")
    buf.seek(0)
    return buf


def _to_parquet_buffer(raw_bytes: bytes, key: str) -> io.BytesIO:
    """Detect file type by S3 key extension and convert to Parquet."""
    if key.lower().endswith((".xlsx", ".xls")):
        return _xlsx_bytes_to_parquet_buffer(raw_bytes)
    return _csv_bytes_to_parquet_buffer(raw_bytes)


def ingest_trust_provider_ref_to_bronze(
    *,
    s3: Any,
    trust_bucket: str,
    trust_key: str,
    platform_bucket: str,
    ingest_date: date,
    env: str,
    source_name: str = "trust_s3_provider_ref",
    kms_key_arn: str | None = None,
) -> dict[str, Any]:
    run_id = str(uuid.uuid4())
    started_at = utc_now_iso()

    bound_log = log.bind(run_id=run_id, source=source_name, env=env)

    manifest_prefix = build_manifest_prefix(source=source_name, ingest_date=ingest_date.isoformat())

    if should_skip_if_already_successful(
        s3=s3, bucket=platform_bucket, manifest_prefix=manifest_prefix
    ):
        bound_log.info("ingest_skipped", reason="latest_manifest_success")
        return {
            "source": source_name,
            "run_id": run_id,
            "env": env,
            "ingest_date": ingest_date.isoformat(),
            "status": "skipped",
            "reason": "latest_manifest_success",
        }

    bronze_key = (
        f"bronze/source={source_name}/entity=provider_site_reference/"
        f"ingest_date={ingest_date.isoformat()}/run_id={run_id}/provider_site_reference.parquet"
    )

    response = s3.get_object(Bucket=trust_bucket, Key=trust_key)
    raw_bytes = response["Body"].read()
    parquet_buf = _to_parquet_buffer(raw_bytes, trust_key)
    extra = s3_kms_args(kms_key_arn)
    s3.upload_fileobj(
        Fileobj=parquet_buf,
        Bucket=platform_bucket,
        Key=bronze_key,
        ExtraArgs=extra if extra else None,
    )

    finished_at = utc_now_iso()

    manifest = Manifest(
        source=source_name,
        env=env,
        run_id=run_id,
        ingest_date=ingest_date.isoformat(),
        started_at=started_at,
        finished_at=finished_at,
        status="success",
        inputs={"trust_bucket": trust_bucket, "trust_key": trust_key},
        outputs={
            "objects_written": 1,
            "objects": [{"trust_key": trust_key, "s3_key": bronze_key}],
        },
    )

    write_manifest(s3=s3, bucket=platform_bucket, manifest=manifest, kms_key_arn=kms_key_arn)
    bound_log.info("ingest_done", status="success")
    return manifest.model_dump()


def _discover_export_dates(s3: Any, trust_bucket: str, prefix_root: str) -> list[date]:
    """List available export_date partitions under a Trust S3 prefix."""
    prefix_root = prefix_root.rstrip("/") + "/"
    paginator = s3.get_paginator("list_objects_v2")
    dates: set[date] = set()
    for page in paginator.paginate(Bucket=trust_bucket, Prefix=prefix_root, Delimiter="/"):
        for cp in page.get("CommonPrefixes", []):
            m = _EXPORT_DATE_RE.search(cp["Prefix"])
            if m:
                raw = m.group(1)
                dates.add(date(int(raw[:4]), int(raw[4:6]), int(raw[6:8])))
    return sorted(dates)


def ingest_trust_diagnostics_export_date_to_bronze(
    *,
    s3: Any,
    trust_bucket: str,
    prefix_root: str,
    export_date: date | None = None,
    platform_bucket: str,
    env: str,
    source_name: str = "trust_s3_diagnostics",
    fail_fast: bool = True,
    kms_key_arn: str | None = None,
) -> dict[str, Any]:
    if export_date is None:
        available = _discover_export_dates(s3, trust_bucket, prefix_root)
        if not available:
            log.info("trust_s3_no_export_dates", prefix_root=prefix_root)
            return {
                "source": source_name,
                "env": env,
                "run_id": str(uuid.uuid4()),
                "status": "skipped",
                "reason": "no_export_dates_found",
            }
        log.info(
            "trust_s3_discovered_dates",
            count=len(available),
            dates=[d.isoformat() for d in available],
        )
        last_result: dict[str, Any] = {}
        for d in available:
            last_result = ingest_trust_diagnostics_export_date_to_bronze(
                s3=s3,
                trust_bucket=trust_bucket,
                prefix_root=prefix_root,
                export_date=d,
                platform_bucket=platform_bucket,
                env=env,
                source_name=source_name,
                fail_fast=fail_fast,
                kms_key_arn=kms_key_arn,
            )
            if last_result.get("status") == "failed" and fail_fast:
                return last_result
        return last_result

    run_id = str(uuid.uuid4())
    started_at = utc_now_iso()

    bound_log = log.bind(run_id=run_id, source=source_name, env=env)

    manifest_prefix = build_manifest_prefix(source=source_name, ingest_date=export_date.isoformat())

    if should_skip_if_already_successful(
        s3=s3, bucket=platform_bucket, manifest_prefix=manifest_prefix
    ):
        bound_log.info("ingest_skipped", reason="latest_manifest_success")
        return {
            "source": source_name,
            "env": env,
            "ingest_date": export_date.isoformat(),
            "run_id": run_id,
            "status": "skipped",
            "reason": "latest_manifest_success",
        }

    prefix_root = prefix_root.rstrip("/")
    trust_prefix = f"{prefix_root}/export_date={export_date.isoformat().replace('-', '')}/"

    results: list[dict[str, Any]] = []
    status: ManifestStatus = "success"
    run_errors: list[str] = []

    paginator = s3.get_paginator("list_objects_v2")

    for page in paginator.paginate(Bucket=trust_bucket, Prefix=trust_prefix):
        for obj in page.get("Contents", []):
            trust_key = obj.get("Key")
            if not trust_key:
                continue
            if not trust_key.lower().endswith(".csv"):
                continue
            try:
                src_filename = trust_key.split("/")[-1] or "part.csv"
                # Derive parquet filename
                parquet_filename = (
                    src_filename[:-4] + ".parquet"
                    if src_filename.lower().endswith(".csv")
                    else src_filename + ".parquet"
                )
                bronze_key = (
                    f"bronze/source={source_name}/entity=diagnostics_orders/"
                    f"ingest_date={export_date.isoformat()}/run_id={run_id}/{parquet_filename}"
                )

                response = s3.get_object(Bucket=trust_bucket, Key=trust_key)
                raw_bytes = response["Body"].read()
                parquet_buf = _csv_bytes_to_parquet_buffer(raw_bytes)
                extra = s3_kms_args(kms_key_arn)
                s3.upload_fileobj(
                    Fileobj=parquet_buf,
                    Bucket=platform_bucket,
                    Key=bronze_key,
                    ExtraArgs=extra if extra else None,
                )

                results.append(
                    {
                        "trust_key": trust_key,
                        "s3_key": bronze_key,
                        "bytes": int(obj.get("Size", 0)),
                        "etag": obj.get("ETag"),
                        "status": "success",
                    }
                )
                bound_log.info("object_uploaded", trust_key=trust_key, bronze_key=bronze_key)
            except Exception as e:
                status = "failed"
                err = f"{type(e).__name__}: {e}"
                run_errors.append(err)
                results.append({"trust_key": trust_key, "status": "failed", "error": err})
                bound_log.error("object_upload_failed", trust_key=trust_key, error=err)
                if fail_fast:
                    break
        if status == "failed" and fail_fast:
            break

    if not results:
        bound_log.info(
            "trust_s3_no_objects",
            trust_bucket=trust_bucket,
            trust_prefix=trust_prefix,
        )
        finished_at = utc_now_iso()
        manifest = Manifest(
            source=source_name,
            env=env,
            run_id=run_id,
            ingest_date=export_date.isoformat(),
            started_at=started_at,
            finished_at=finished_at,
            status="skipped",
            reason="empty_trust_prefix",
            inputs={"trust_bucket": trust_bucket, "trust_prefix": trust_prefix},
            outputs={"objects_written": 0, "objects_failed": 0, "objects": []},
        )
        write_manifest(s3=s3, bucket=platform_bucket, manifest=manifest, kms_key_arn=kms_key_arn)
        return manifest.model_dump()

    finished_at = utc_now_iso()

    manifest = Manifest(
        source=source_name,
        env=env,
        run_id=run_id,
        ingest_date=export_date.isoformat(),
        started_at=started_at,
        finished_at=finished_at,
        status=status,
        error=run_errors,
        inputs={"trust_bucket": trust_bucket, "trust_prefix": trust_prefix},
        outputs={
            "objects_written": sum(1 for r in results if r.get("status") == "success"),
            "objects_failed": sum(1 for r in results if r.get("status") == "failed"),
            "objects": results,
        },
    )

    write_manifest(s3=s3, bucket=platform_bucket, manifest=manifest, kms_key_arn=kms_key_arn)
    bound_log.info("ingest_done", status=status)
    return manifest.model_dump()
